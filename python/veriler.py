# Bu kodlar eldeki df'i incelemeye uygun hale getiriyor. gerekli indicatörleri ekliyor flaan filan...
# İstenen veriyi en uygun şekilde hazırlıyor.
import MetaTrader5 as mt5
import ta
import openpyxl
import pandas as pd

def destek_mi(df1,l,n1,n2):
    for j in range(l-n1+1,l+1):
        if df1.low[j]>df1.low[j-1]:
            return 0
    for j in range(l+1,l+n2+1):
        if df1.low[j]<df1.low[j-1]:
            return 0
    return 1

class Veri:
    def __init__(self):
        pass

    def veri_al(self, ürün_adı, periyot, uzaklık):
        mt5.initialize()
        if periyot == "1M":
            veri1 = mt5.copy_rates_from_pos(
                ürün_adı, mt5.TIMEFRAME_M1, 0, uzaklık)
            df = pd.DataFrame(veri1)
        elif periyot == "2M":
            veri1 = mt5.copy_rates_from_pos(
                ürün_adı, mt5.TIMEFRAME_M2, 0, uzaklık)
            df = pd.DataFrame(veri1)
        elif periyot == "3M":
            veri1 = mt5.copy_rates_from_pos(
                ürün_adı, mt5.TIMEFRAME_M3, 0, uzaklık)
            df = pd.DataFrame(veri1)
        elif periyot == "5M":
            veri1 = mt5.copy_rates_from_pos(
                ürün_adı, mt5.TIMEFRAME_M5, 0, uzaklık)
            df = pd.DataFrame(veri1)
        elif periyot == "10M":
            veri1 = mt5.copy_rates_from_pos(
                ürün_adı, mt5.TIMEFRAME_M10, 0, uzaklık)
            df = pd.DataFrame(veri1)
        elif periyot == "15M":
            veri1 = mt5.copy_rates_from_pos(
                ürün_adı, mt5.TIMEFRAME_M15, 0, uzaklık)
            df = pd.DataFrame(veri1)
        elif periyot == "30M":
            veri1 = mt5.copy_rates_from_pos(
                ürün_adı, mt5.TIMEFRAME_M30, 0, uzaklık)
            df = pd.DataFrame(veri1)
        elif periyot == "1H":
            veri1 = mt5.copy_rates_from_pos(
                ürün_adı, mt5.TIMEFRAME_H1, 0, uzaklık)
            df = pd.DataFrame(veri1)
        elif periyot == "2H":
            veri1 = mt5.copy_rates_from_pos(
                ürün_adı, mt5.TIMEFRAME_H2, 0, uzaklık)
            df = pd.DataFrame(veri1)
        elif periyot == "4H":
            veri1 = mt5.copy_rates_from_pos(
                ürün_adı, mt5.TIMEFRAME_H4, 0, uzaklık)
            df = pd.DataFrame(veri1)
        elif periyot == "6H":
            veri1 = mt5.copy_rates_from_pos(
                ürün_adı, mt5.TIMEFRAME_H6, 0, uzaklık)
            df = pd.DataFrame(veri1)
        elif periyot == "12H":
            veri1 = mt5.copy_rates_from_pos(
                ürün_adı, mt5.TIMEFRAME_H12, 0, uzaklık)
            df = pd.DataFrame(veri1)
        elif periyot == "1D":
            veri1 = mt5.copy_rates_from_pos(
                ürün_adı, mt5.TIMEFRAME_D1, 0, uzaklık)
            df = pd.DataFrame(veri1)
        elif periyot == "1W":
            veri1 = mt5.copy_rates_from_pos(
                ürün_adı, mt5.TIMEFRAME_W1, 0, uzaklık)
            df = pd.DataFrame(veri1)
        df = df.drop(df.columns[[6, 7]], axis=1)
        df["time"] = pd.to_datetime(df["time"], unit="s")
        return (df)

    def veri_indikatör_ekle_EMA(self, df, periyot):
        df["EMA"+str(periyot)] = ta.trend.EMAIndicator(close=df.close,
                                                       window=periyot, fillna=False).ema_indicator()

    def veri_indikatör_ekle_ATR(self, df, periyot):
        df["ATR"] = ta.volatility.AverageTrueRange(
            df.high, df.low, df.close, window=periyot, fillna=1).average_true_range()

    def veri_indikatör_ekle_RSI(self, df, periyot):       
        df["RSI"]=ta.momentum.RSIIndicator(close=df.close,
                                                       window=periyot, fillna=False).rsi() 

    def ortalama_kesişimi_veri(self,df,hızlı_ad,yavaş_ad,column_name):
        df[column_name]= abs(df[hızlı_ad]-df[yavaş_ad])/(df[hızlı_ad]-df[yavaş_ad])
        
    def pivot_destekler(self,df,soldan:int,sağdan:int):
        index_ve_destek_listesi = []
        for i in range(soldan+1,len(df)-sağdan):
            if destek_mi(df,i,soldan,sağdan):
                index_ve_destek_listesi.append((i,df.loc[i].low))
        return index_ve_destek_listesi

    def geçmiş_işlemleri_al(self,ürün_adı:str,):
        pass

    # Şimdilik lot ayarlayamıyor,komisyonu 2defa sayıyor başka sorun gözükmüyor şimdilik
    def metatrader_backtest_verisi_al(self,dosya_adı:str):
        # Give the location of the file
        path = "C:\\Users\\Bayoglu\\Desktop\\Day Trade\\"+dosya_adı+".xlsx"

        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active
        starts_of_positions = int(((sheet_obj.max_row-55)/2)+58)
        cell_obj = sheet_obj.cell(row=starts_of_positions, column=1)

        işlem_sayısı = int(sheet_obj.cell(row=27, column=4).value)

        # Şimdi alacağım yerin başlangıcını buldum burdan sona kadar gideceğim
        işlemler = []
        # Row cinsi
        for i in range(starts_of_positions, starts_of_positions+işlem_sayısı*2, 2):
            işleme_giriş_tarihi = sheet_obj.cell(row=i, column=1).value
            işleme_giriş_fiyatı = float(sheet_obj.cell(row=i, column=7).value)
            işlem_yönü = sheet_obj.cell(row=i, column=4).value
            if işlem_yönü == "buy":
                işlem_yönü = "Buy"
            else:
                işlem_yönü = "Sell"
            işlem_hacmi = sheet_obj.cell(row=i, column=6).value
            işlemin_kârı = sheet_obj.cell(row=i+1, column=11).value
            işlem_sonu_bakiye = sheet_obj.cell(row=i+1, column=12).value
            işlemden_çıkış_fiyatı = float(sheet_obj.cell(row=i + 1, column=7).value)
            işlemden_çıkış_tarihi = sheet_obj.cell(row=i+1, column=1).value
            # Sadece işlemleri alacak o yüzdne giriş çıkış endexlerle tp sl ler yok
            işlemler.append((işlem_yönü,işleme_giriş_tarihi,işleme_giriş_fiyatı,işlemden_çıkış_tarihi,işlemden_çıkış_fiyatı,0,0,0,0))
        return işlemler
        



       
        
        






